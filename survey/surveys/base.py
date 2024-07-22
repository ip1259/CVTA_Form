import datetime
import json
import os

import gradio as gr
import openpyxl
from gradio.themes.base import Base as grBase

import survey.themes.themes as s_themes
from survey.pages import *
from survey.exceptions import UnableLoadSurvey


class BaseSurvey:
    def __init__(self):
        self.survey_id: str = ""
        self.survey_theme: tuple[grBase, str] | None = None
        self.heads: list[Page] = []
        self.bodies: list[Page] = []
        self.tail: Page | None = None
        self.survey: gr.Blocks | None = None
        self.user_store: gr.State | None = None

    def get_rows_in_bodies(self):
        _results = []
        _results.extend([b.page for b in self.bodies])
        return _results

    def get_survey_input_components(self):
        _results = []
        for p in self.bodies:
            _results.append(p.get_input_components())
        return _results

    def set_survey(self, file_path: str):
        """
        透過描述問卷的json檔案設定Survey物件
        Survey File Structure:
        {
        'survey_id': 字串-問卷的ID, 'survey_theme': "EMB"-預設值,
        'heads': [{
            'page_type': 'head',
            'title': "OOO班 教師滿意度調查",
            'desc': 字串-可選，描述性說明文字
            }],
        'bodies': [
            {'page_type': 'teacher', 'teacher': 字串-必須有-老師名稱, 'max_score': 整數-可選，最高分},
            {'page_type': 'ta', 'max_score': 整數-可選，最高分},
            {'page_type': 'employment'，就業輔導評分},
            {'page_type': "finish"，最後必需有，問卷送出後的顯示頁面},
            {
            'page_type': "custom"，自定義組合,
            'block_dicts':[
                {'block_type': "head", 'title': 字串，必須，標題, 'desc': 字串，必須，說明文字},頁首標題方塊
                {'block_type': "finish", 'title': 字串，可選，結束語},顯示結束的純文字方塊
                {
                'block_type': "score",-----評分問題
                'title': 字串，必須，描述問題,
                'must': 布林值，可選，設定問題是否為必填,
                'max_score': 整數值，可選，設定問題最高分,
                'max_score_desc': 字串，可選，設定最高分顯示標籤,
                'min_score_desc': 字串，可選，設定最低分顯示標籤},
                {
                'block_type': "suggestion",-----建議
                'title': 字串，必須，描述問題,
                'must': 布林值，可選，設定問題是否為必填
                }]
            }],
        'tail': {'page_type': 'tail_btn'必需，底部按鈕列, 'body_count': int，不須設置，由系統自動取得}}
        :param file_path:描述問卷的json檔案位置
        :return: None
        """
        with open(file_path, 'r', encoding='utf8') as _sf:
            _survey_dict = json.load(_sf)
            self.load(_survey_dict)

    def on_survey_load(self, request: gr.Request):
        pass

    def on_survey_unload(self, request: gr.Request):
        pass

    def next_click(self, data):
        """

        :param data: data應為以component本身為鍵值的字典,[0]應為gradio.State的value值，為一字典包含
        :return:
        """
        print(data)
        _user_store_key = None
        for _k in data.keys():
            if isinstance(_k, gr.State):
                _user_store_key = _k
                break
        _cur: int = data[_user_store_key]['cur_page']
        _t_list = [(True if ((not _m) or (data[_c] is not None)) else False) for _m, _c, _t in
                   self.bodies[_cur].get_input_components()]
        _must_all_done = all([(True if ((not _m) or (data[_c] is not None)) else False) for _m, _c, _t in
                              self.bodies[_cur].get_input_components()])
        if _must_all_done:
            _cur = data[_user_store_key]['cur_page'] + 1
            _cur = min(_cur, len(self.bodies) - 1)
            _cur = max(_cur, 0)
            data[_user_store_key]['cur_page'] = _cur
        return data[_user_store_key]

    def prev_click(self, data):
        """

        :param data: data應為以component本身為鍵值的字典,[0]應為gradio.State的value值，為一字典包含
        :return:
        """

        _user_store_key = None
        for _k in data.keys():
            if isinstance(_k, gr.State):
                _user_store_key = _k
                break
        _cur: int = data[_user_store_key]['cur_page']
        _cur = data[_user_store_key]['cur_page'] - 1
        _cur = min(_cur, len(self.bodies) - 1)
        _cur = max(_cur, 0)
        data[_user_store_key]['cur_page'] = _cur
        return data[_user_store_key]

    def send_click(self, data, request: gr.Request):
        """

        :param request:
        :param data: data應為以component本身為鍵值的字典,[0]應為gradio.State的value值，為一字典包含"cur_page", "results"兩鍵值
        "cur_page為目前body的索引值", "results"為一list依序存放第一題到最後一題的tuple(問題字串, value)
        :return:
        """

        _user_store_key = None
        for _k in data.keys():
            if isinstance(_k, gr.State):
                _user_store_key = _k
                break
        _cur: int = data[_user_store_key]['cur_page']
        _must_all_done = all([(True if ((not _m) or (data[_c] is not None)) else False) for _m, _c, _t in
                              self.bodies[_cur].get_input_components()])
        if _must_all_done:
            _cur = data[_user_store_key]['cur_page'] + 1
            _cur = min(_cur, len(self.bodies) - 1)
            _cur = max(_cur, 0)
            data[_user_store_key]['cur_page'] = _cur
            for _i, _b in enumerate(self.bodies):
                if len(_b.get_input_components()) == 0:
                    data[_user_store_key]['results'].pop(_i)
                for _j, (_m, _c, _t) in enumerate(_b.get_input_components()):
                    data[_user_store_key]['results'][_i][_j] = (_t, data[_c])
        if request:
            _ip = request.request.client.host
            data[_user_store_key]['ip'] = _ip
        data[_user_store_key]["time_stamp"] = datetime.datetime.now().isoformat(" ", "minutes")
        self.save_user_store(data[_user_store_key])
        return data[_user_store_key]

    def user_store_changed(self, _user_store):
        print(_user_store)
        _cur = _user_store['cur_page']
        _cur_page: gr.Row = self.bodies[_cur].page
        _output = {}
        if isinstance(self.tail, TailButtonPage):
            _output[self.tail.page_blocks[0].prev_warp] = (
                self.tail.page_blocks[0].prev_warp.update(_cur > 0 and _cur != len(self.bodies) - 1))
            _output[self.tail.page_blocks[0].next_warp] = (
                self.tail.page_blocks[0].next_warp.update(_cur < len(self.bodies) - 2))
            _output[self.tail.page_blocks[0].send_warp] = (
                self.tail.page_blocks[0].send_warp.update(_cur == len(self.bodies) - 2))
        for _i, _r in enumerate(self.bodies):
            _output[_r.page] = _r.page.update(True if _i == _cur else False)
        return _output

    def load(self, survey_dict: dict):
        if all(_i in survey_dict.keys() for _i in ['survey_id', 'survey_theme', 'heads', 'bodies', 'tail']):
            try:
                self.survey_id = survey_dict['survey_id']
                self.set_theme(survey_dict['survey_theme'])

                with (gr.Blocks(theme=self.survey_theme[0], css=self.survey_theme[1],
                                js=s_themes.SurveyTheme.JS) as self.survey):
                    _user_data = {"cur_page": 0, "ip": "", "time_stamp": "", "total_score": 0, "results": []}
                    self.user_store = gr.State()
                    with gr.Row():
                        gr.Column(scale=1, min_width=0)
                        with gr.Column(scale=3, min_width=640):
                            for _hp in survey_dict['heads']:
                                _t = PageParser.parse_page(_hp)
                                self.heads.append(_t)
                                _t.set_page_interactive()
                            for _i, _bp in enumerate(survey_dict['bodies']):
                                _b = PageParser.parse_page(_bp)
                                _user_data['results'].append([None for _i in range(len(_b.get_input_components()))])
                                self.bodies.append(_b)
                                _b.set_page_interactive()
                                if len(self.bodies) == 1:
                                    _b.page.visible = True
                                else:
                                    _b.page.visible = False
                                if isinstance(_b, FinishPage):
                                    break
                                elif _i == len(survey_dict['bodies']) - 1:
                                    _b = FinishPage()
                                    self.bodies.append(_b)
                                    _b.set_page_interactive()

                            self.tail = TailButtonPage(len(self.bodies))
                            _btns = self.tail.get_buttons()
                            self.user_store.value = _user_data
                            _inputs = [self.user_store]
                            [_inputs.extend([_t[1] for _t in _p]) for _p in self.get_survey_input_components()]
                            _inputs_set = set(_inputs)
                            _outputs = [_p.page for _p in self.bodies]
                            _outputs.extend([self.tail.page_blocks[0].prev_warp,
                                             self.tail.page_blocks[0].next_warp,
                                             self.tail.page_blocks[0].send_warp])
                            _outputs_set = set(_outputs)

                            to_top_js = """
                                        function to_top() {
                                            window.scrollTo(0,0);
                                        }
                                        """

                            _btns[0].click(self.prev_click, inputs=_inputs_set, outputs={self.user_store})
                            _btns[1].click(self.next_click, inputs=_inputs_set, outputs=self.user_store)
                            _btns[2].click(self.send_click, inputs=_inputs_set, outputs={self.user_store})
                            self.user_store.change(self.user_store_changed, inputs=self.user_store,
                                                   outputs=_outputs_set,
                                                   js=to_top_js)

                            # self.survey.load(self.on_survey_load)
                            # self.survey.unload(self.on_survey_unload)
                            self.user_store = gr.State(_user_data)

                        gr.Column(scale=1, min_width=0)

            except Exception as e:
                print(e)
                raise UnableLoadSurvey(-1)
        else:
            raise UnableLoadSurvey(0)

    def set_theme(self, theme: str):
        match theme:
            case "EMB":
                self.survey_theme = s_themes.SurveyTheme.EMB
            case _:
                self.survey_theme = s_themes.SurveyTheme.EMB

    def save_user_store(self, _user_store: dict):
        _path = os.getcwd()
        _path = os.path.join(_path, "responses")
        if not os.path.exists(_path):
            os.mkdir(_path)
        _path = os.path.join(_path, self.survey_id)
        if not os.path.exists(_path):
            os.mkdir(_path)
        from joblib import dump
        dump(_user_store, os.path.join(_path, _user_store['ip'].split('.')[-1] + ".rep"))

    def save_to_excel(self, _user_store: dict):
        _path = os.getcwd()
        _path = os.path.join(_path, "responses")
        if not os.path.exists(_path):
            os.mkdir(_path)
        _path = os.path.join(_path, self.survey_id)
        if not os.path.exists(_path):
            os.mkdir(_path)
        _wb = None
        _path = os.path.join(_path, self.survey_id + ".xlsx")
        if not os.path.exists(_path):
            _wb = openpyxl.Workbook()
            _wb.create_sheet("表單回覆")
            _wb.remove_sheet(_wb["Sheet"])
            _ws = _wb["表單回覆"]
            _titles = ["時間戳記", "分數", "IP"]
            [_titles.extend([_q[0] for _q in _p]) for _p in _user_store['results']]
            _ws.append(_titles)
            _wb.save(_path)
        _wb = openpyxl.load_workbook(_path)
        _response = [_user_store['time_stamp'], _user_store['total_score'], _user_store['ip']]
        [_response.extend([_q[1] for _q in _p]) for _p in _user_store['results']]
        for i, r in enumerate(_response):
            if r is None:
                _response[i] = ""
        _ws = _wb["表單回覆"]
        _ws.append(_response)
        _wb.save(_path)

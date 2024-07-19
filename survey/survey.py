import os.path

import gradio.components.base
from gradio.themes.base import Base

import survey.blocks.input_block
from survey.pages.page import Page
from survey.pages.pages import *
from survey.themes.themes import *

import openpyxl


class Survey:
    def __init__(self, class_name: str, teachers: list[str],
                 survey_id: str,
                 has_ta: bool,
                 survey_desc: str = "本問卷用以了解學生對於師資滿意度進行調查，以匿名方式提供老師日後教學上的參考，請放心填寫。\n*代表必填",
                 survey_theme: Base = SurveyTheme.EMB):
        self.has_ta = has_ta
        self.survey_id = survey_id
        self.survey_theme = survey_theme
        self.survey_desc = survey_desc
        self.teachers = teachers
        self.class_name = class_name
        self.head: Page | None = None
        self.body: list[Page] | None = []
        self.body_rows: list[gradio.components.base.Component] = []
        self.clients: dict[str, SurveyClient] = {}
        self.response: dict[str, SurveyResponse] | None = {}

    def start_survey(self):
        with gr.Blocks(js=SurveyTheme.JS, css=SurveyTheme.EMB_CSS, theme=self.survey_theme) as _survey:
            _body_rows: list[gradio.components.base.Component] = []
            self.head = HeadPage(self.class_name + "教師滿意度調查問卷", self.survey_desc, self)
            for i, teacher in enumerate(self.teachers):
                with gr.Row(visible=False) as r:
                    _t = TeacherPage(teacher, self)
                    self.body.append(_t)
                _body_rows.append(r)
                if i == 0:
                    r.visible = True
            with gr.Row(visible=False) as emp_row:
                _e = EmploymentPage(self)
                self.body.append(_e)
                _body_rows.append(emp_row)
            if self.has_ta:
                with gr.Row(visible=False) as ta_row:
                    _ta = TAPage(self)
                    self.body.append(_ta)
                    _body_rows.append(ta_row)

            _fin = FinishPage(self)
            _fin.page.visible = False
            self.body.append(_fin)
            _body_rows.append(_fin.page)

            self.body_rows = _body_rows

            with gr.Row():
                with gr.Column(min_width=0, scale=1):
                    pass
                with gr.Column(variant="default", scale=3, min_width=640):
                    with gr.Row():
                        with gr.Row(visible=False) as prev_warp:
                            prev_button = gr.Button("返回", min_width=75, scale=1)
                        with gr.Column(scale=10, min_width=0):
                            pass
                        with gr.Row(visible=True) as next_warp:
                            next_button = gr.Button("繼續", min_width=75, scale=1)
                        with gr.Row(visible=False) as send_warp:
                            send_button = gr.Button("送出", min_width=75, scale=1, elem_id="sendButton")

                with gr.Column(min_width=0, scale=1):
                    pass

            def next_body(request: gr.Request):
                if request:
                    _ip = request.client.host
                    _cur = self.clients[_ip].cur_body_index
                    if self.body[_cur].must_has_done(_ip):
                        self.clients[_ip].set_cur_body(_cur + 1)
                    _cur = self.clients[_ip].cur_body_index
                    _result = []
                    for _i, _r in enumerate(_body_rows):
                        if isinstance(_r, gr.Row):
                            _result.append(_r.update(self.clients[_ip].body_visible_states[_i]))
                    _result.extend([prev_warp.update(_cur > 0 and _cur != len(
                        _body_rows) - 1),
                                    next_warp.update((_cur < len(
                                        _body_rows) - 2)),
                                    send_warp.update((_cur == len(
                                        _body_rows) - 2))])
                    return _result
                return None

            def prev_body(request: gr.Request):
                if request:
                    _ip = request.client.host
                    _cur = self.clients[_ip].cur_body_index
                    self.clients[_ip].set_cur_body(_cur - 1)
                    _cur = self.clients[_ip].cur_body_index
                    _result = []
                    for _i, _r in enumerate(_body_rows):
                        if isinstance(_r, gr.Row):
                            _result.append(_r.update(self.clients[_ip].body_visible_states[_i]))
                    _result.extend([prev_warp.update(_cur > 0 and _cur != len(
                        _body_rows) - 1),
                                    next_warp.update((_cur < len(
                                        _body_rows) - 2)),
                                    send_warp.update((_cur == len(
                                        _body_rows) - 2))])
                    return _result
                return None

            def send_click(request: gr.Request):
                if request:
                    _ip = request.request.client.host
                    print(self.clients[_ip].response)
                    self.clients[_ip].save_response()

                    _cur = self.clients[_ip].cur_body_index
                    if self.body[_cur].must_has_done(_ip):
                        self.clients[_ip].set_cur_body(_cur + 1)
                    _cur = self.clients[_ip].cur_body_index
                    _result = []
                    for _i, _r in enumerate(_body_rows):
                        if isinstance(_r, gr.Row):
                            _result.append(_r.update(self.clients[_ip].body_visible_states[_i]))
                    _result.extend([prev_warp.update(_cur > 0 and _cur != len(
                        _body_rows) - 1),
                                    next_warp.update((_cur < len(
                                        _body_rows) - 2)),
                                    send_warp.update((_cur == len(
                                        _body_rows) - 2))])
                    return _result
                return None

            btn_outputs = [_r for _r in _body_rows]
            btn_outputs.extend([prev_warp, next_warp, send_warp])
            to_top_js = """
                        function to_top() {
                            window.scrollTo(0,0);
                        }
                        """
            next_button.click(next_body, None,
                              outputs=btn_outputs, js=to_top_js)
            prev_button.click(prev_body, None,
                              outputs=btn_outputs, js=to_top_js)
            send_button.click(send_click, None,
                              outputs=btn_outputs, js=to_top_js)

            def survey_load(request: gr.Request):
                _ip = request.request.client.host
                self.clients[_ip] = SurveyClient(_ip, self)

            def survey_unload(request: gr.Request):
                _ip = request.request.client.host
                self.clients.pop(_ip)

            _survey.load(survey_load)
            _survey.unload(survey_unload)
        return _survey

    def get_survey_input_components(self):
        _results = []
        for p in self.body:
            _results.extend(p.get_input_components())
        return _results


class SurveyResponse:
    def __init__(self, parent: Survey):
        self.parent = parent
        # example self.response[an InputBlock Object as KEY] = (Question, Must Answer Or Not, Value)
        self.response: dict[survey.blocks.input_block.InputBlock, tuple[str, bool, str | int | float]] = {}
        self.init_response()

    def init_response(self):
        _inputs = self.parent.get_survey_input_components()
        for i in _inputs:
            self.set_response(i, None)

    def set_response(self, input_block: survey.blocks.input_block.InputBlock, value):
        self.response[input_block] = (input_block.title, input_block.must, value)


class SurveyClient:
    def __init__(self, ip: str, parent: Survey):
        self.ip = ip
        self.parent = parent
        self.response = SurveyResponse(parent)
        self.cur_body_index: int = 0
        self.body_visible_states: list[bool] = [r.visible for r in self.parent.body_rows]

    def save_response(self):
        _path = os.getcwd()
        _path = os.path.join(_path, "responses")
        if not os.path.exists(_path):
            os.mkdir(_path)
        _path = os.path.join(_path, self.parent.survey_id)
        if not os.path.exists(_path):
            os.mkdir(_path)
        from joblib import dump, load
        dump([i for i in self.response.response.values()], os.path.join(_path, self.ip.split('.')[-1]+".rep"))
        # example: list[tuple[str, bool, str | int | float]] = load("[WorkDir]/responses/[SurveyID]/[IPv4最後一位.rep]")
        #
        # _wb = None
        # _path = os.path.join(_path, self.parent.survey_id + ".xlsx")
        # if not os.path.exists(_path):
        #     _wb = openpyxl.Workbook()
        #     _wb.create_sheet("表單回覆")
        #     _wb.remove_sheet(_wb["Sheet"])
        #     _ws = _wb["表單回覆"]
        #     _titles = ["IP"]
        #     _titles.extend([_t[0] for _t in self.response.response.values()])
        #     _ws.append(_titles)
        #     _wb.save(_path)
        # _wb = openpyxl.load_workbook(_path)
        # _response = [self.ip]
        # _response.extend([_t[2] for _t in self.response.response.values()])
        # for i, r in enumerate(_response):
        #     if r is None:
        #         _response[i] = ""
        # _ws = _wb["表單回覆"]
        # _ws.append(_response)
        # _wb.save(_path)

    def set_cur_body(self, index: int):
        self.cur_body_index = index
        self.cur_body_index = min(self.cur_body_index, len(self.body_visible_states) - 1)
        self.cur_body_index = max(self.cur_body_index, 0)
        for _i in range(len(self.body_visible_states)):
            if _i == self.cur_body_index:
                self.body_visible_states[_i] = True
            else:
                self.body_visible_states[_i] = False
